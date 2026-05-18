import streamlit as st
import pandas as pd
import plotly.express as px
from openai import OpenAI
from werkzeug.security import check_password_hash
from pathlib import Path
from io import BytesIO
from copy import copy
from datetime import date
from openpyxl import load_workbook
from openpyxl.formula.translate import Translator
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.page import PageMargins
import json
import os
import re

# =====================================================
# OPENAI
# =====================================================

client = OpenAI(
    api_key=os.getenv("OPENAI_API_KEY")
)

# =====================================================
# PAGE CONFIG
# =====================================================

st.set_page_config(
    page_title="Nursery Intelligence Copilot Enterprise",
    layout="wide"
)

# =====================================================
# LOGIN CONFIG
# =====================================================

BUILTIN_ADMIN_USERNAME = "admin"

BUILTIN_ADMIN_PASSWORD_HASH = (
    "scrypt:32768:8:1$cer4Rc98UopaCXLm$4d12e32c06be3f94fab0e22a80237c5d552f4b85ec0167b4c7f422050af7228161b83c239f4eae0fdc3dcdc17068492cc04dbcbc1a87bd565df57d776993abba"
)

STANDARD_COLUMNS = [
    "Date",
    "Client Name",
    "Crop Name",
    "Variety",
    "Line",
    "Rep",
    "Quantity",
    "Amount",
    "Cavity",
    "Seeds Min",
    "Seeds Max",
    "Seeds Recommended"
]

DATA_CACHE_VERSION = 8


def empty_standard_frame():

    return pd.DataFrame({
        "Date": pd.Series(dtype="datetime64[ns]"),
        "Client Name": pd.Series(dtype="string"),
        "Crop Name": pd.Series(dtype="string"),
        "Variety": pd.Series(dtype="string"),
        "Line": pd.Series(dtype="string"),
        "Rep": pd.Series(dtype="string"),
        "Quantity": pd.Series(dtype="float"),
        "Amount": pd.Series(dtype="float"),
        "Cavity": pd.Series(dtype="float"),
        "Seeds Min": pd.Series(dtype="float"),
        "Seeds Max": pd.Series(dtype="float"),
        "Seeds Recommended": pd.Series(dtype="float")
    })


def ensure_standard_columns(df):

    for col in STANDARD_COLUMNS:

        if col not in df.columns:

            if col == "Date":
                df[col] = pd.NaT

            elif col in [
                "Quantity",
                "Amount",
                "Cavity",
                "Seeds Min",
                "Seeds Max",
                "Seeds Recommended"
            ]:
                df[col] = 0

            else:
                df[col] = ""

    df["Date"] = pd.to_datetime(
        df["Date"],
        errors="coerce"
    )

    df["Quantity"] = pd.to_numeric(
        df["Quantity"],
        errors="coerce"
    ).fillna(0)

    df["Amount"] = pd.to_numeric(
        df["Amount"],
        errors="coerce"
    ).fillna(0)

    df["Cavity"] = pd.to_numeric(
        df["Cavity"],
        errors="coerce"
    ).fillna(0)

    df["Seeds Min"] = pd.to_numeric(
        df["Seeds Min"],
        errors="coerce"
    ).fillna(0)

    df["Seeds Max"] = pd.to_numeric(
        df["Seeds Max"],
        errors="coerce"
    ).fillna(0)

    df["Seeds Recommended"] = pd.to_numeric(
        df["Seeds Recommended"],
        errors="coerce"
    ).fillna(0)

    return df


def parse_excel_dates(series):

    numeric_dates = pd.to_numeric(
        series,
        errors="coerce"
    )

    parsed = pd.to_datetime(
        series,
        errors="coerce"
    )

    excel_serial_mask = (
        numeric_dates.notna()
        & numeric_dates.between(
            20000,
            60000
        )
    )

    if excel_serial_mask.any():

        parsed.loc[excel_serial_mask] = pd.to_datetime(
            numeric_dates.loc[excel_serial_mask],
            unit="D",
            origin="1899-12-30",
            errors="coerce"
        )

    return parsed


def add_order_seed_columns(df):

    d = df.copy()

    cavity = pd.to_numeric(
        d["Cavity"],
        errors="coerce"
    ).fillna(0)

    line = (
        d["Line"]
        .astype(str)
        .str.upper()
    )

    seed_min = pd.Series(
        0,
        index=d.index,
        dtype="float"
    )

    seed_max = pd.Series(
        0,
        index=d.index,
        dtype="float"
    )

    six_pack = (
        cavity.eq(6)
        & line.str.contains(
            "SEEDLINGS",
            na=False
        )
    )

    four_pack = (
        cavity.eq(4)
        & line.str.contains(
            "SEEDLINGS",
            na=False
        )
    )

    twelve_cm = (
        cavity.eq(12)
        | line.str.contains(
            "12CM",
            na=False
        )
    )

    fifteen_cm = (
        cavity.eq(15)
        | line.str.contains(
            "15CM COLOUR",
            na=False
        )
    )

    plant_to_plate = (
        cavity.eq(17)
        | line.str.contains(
            "PLANT TO PLATE",
            na=False
        )
    )

    seed_min.loc[six_pack] = 6
    seed_max.loc[six_pack] = 6

    seed_min.loc[four_pack] = 4
    seed_max.loc[four_pack] = 4

    seed_min.loc[twelve_cm] = 1
    seed_max.loc[twelve_cm] = 1

    seed_min.loc[fifteen_cm] = 2
    seed_max.loc[fifteen_cm] = 2

    seed_min.loc[plant_to_plate] = 2
    seed_max.loc[plant_to_plate] = 3

    d["Seeds Min"] = (
        d["Quantity"]
        * seed_min
    )

    d["Seeds Max"] = (
        d["Quantity"]
        * seed_max
    )

    d["Seeds Recommended"] = d["Seeds Max"]

    return d

if "authenticated" not in st.session_state:
    st.session_state.authenticated = False


def user_password_hashes():

    try:

        u = st.secrets.get("users")

        if u is None:
            return {}

        if hasattr(u, "to_dict"):
            u = u.to_dict()

        elif not isinstance(u, dict):
            u = dict(u)

        return {
            str(k): str(v)
            for k, v in u.items()
        }

    except Exception:
        return {}


def verify_login(username, password):

    uname = (username or "").strip()

    if not uname:
        return False

    users = user_password_hashes()

    h = users.get(uname)

    if h and check_password_hash(h, password):
        return True

    if (
        uname.lower() == BUILTIN_ADMIN_USERNAME.lower()
        and check_password_hash(
            BUILTIN_ADMIN_PASSWORD_HASH,
            password
        )
    ):
        return True

    return False


if not st.session_state.authenticated:

    st.title("🌱 Nursery Intelligence Copilot")

    with st.form("login_form"):

        username = st.text_input("Username")

        password = st.text_input(
            "Password",
            type="password"
        )

        submitted = st.form_submit_button(
            "Sign In"
        )

        if submitted:

            if verify_login(
                username,
                password
            ):

                st.session_state.authenticated = True
                st.rerun()

            else:

                st.error(
                    "Invalid username or password"
                )

    st.stop()

# =====================================================
# KNOWN LINES
# =====================================================

KNOWN_LINES = [
    "SEEDLINGS",
    "15CM COLOUR POTS",
    "12CM COLOUR POT",
    "12CM HERB/VEG",
    "PETUNIA HYBRIDS",
    "SIMPLY BEAUTIFUL",
    "PLANT TO PLATE",
    "MID RANGE",
    "CALIBRACHOA",
    "20CM AERO BOWL",
    "14CM SUCCULENTS",
    "25CM HANGING BASKET",
    "25CM COLOUR POT",
    "ARGYRANTHEMUM",
    "15CM RANUNCULUS",
    "15CM ANGELONIA",
    "12CM PATIO RANGE",
    "12CM ORNAMENTAL CHILLI",
    "9CM SUCCULENTS",
    "DAHLIA POTS",
    "12CM PETUNIA HYBRIDS",
    "32CM MIX AND MINGLE",
    "12CM PRIMROSE/OBCONICA",
    "CHILLI POT",
    "LAWN PLUGS",
    "12CM GRASS POTS",
    "17CM GERANIUM"
]

PETUNIA_SEEDABLE_LINES = [
    "SEEDLINGS",
    "12CM COLOUR POT",
    "15CM COLOUR POTS"
]

SOWING_PLANNER_TEMPLATE = "AT_Nursery_Weekly_Sowing_Planner_V3.xlsx"
GROW_WEEKS_FILE = "wholesale_nursery_seedling_sale_weeks.xlsx"
SEED_STOCK_FILE = "seed_stock.xlsx"
SEED_STOCK_FILES = [
    SEED_STOCK_FILE,
    "seed_stock.csv",
    "2026 - Seed Stock.xlsx",
    "2026 - Seed Stock.csv"
]

# =====================================================
# APP DIRECTORY
# =====================================================

def app_dir():

    return Path(__file__).resolve().parent


def template_path():

    local_path = app_dir() / SOWING_PLANNER_TEMPLATE

    if local_path.is_file():
        return local_path

    documents_path = (
        Path.home()
        / "Documents"
        / SOWING_PLANNER_TEMPLATE
    )

    if documents_path.is_file():
        return documents_path

    return local_path


def grow_weeks_path():

    candidates = [
        app_dir() / GROW_WEEKS_FILE,
        Path.home() / "Documents" / GROW_WEEKS_FILE,
        Path.home() / "Downloads" / GROW_WEEKS_FILE
    ]

    for candidate in candidates:

        if candidate.is_file():
            return candidate

    return candidates[0]


def seed_stock_path():

    search_dirs = [
        app_dir(),
        Path.home() / "Documents",
        Path.home() / "Downloads",
        Path("Z:/Seeds & Sowing/Seed Stock")
    ]

    candidates = [
        folder / filename
        for folder in search_dirs
        for filename in SEED_STOCK_FILES
    ]

    for candidate in candidates:

        if candidate.is_file():
            return candidate

    return candidates[0]

# =====================================================
# COLUMN FINDER
# =====================================================

def first_column(df, candidates):

    cols = {
        str(c).strip().lower(): c
        for c in df.columns
    }

    for cand in candidates:

        key = cand.lower().strip()

        if key in cols:
            return cols[key]

    for c in df.columns:

        cl = str(c).strip().lower()

        for cand in candidates:

            if cand.lower() in cl:
                return c

    return None

# =====================================================
# LOAD EXCEL FILE
# =====================================================

def load_excel_file(path):

    candidate = app_dir() / path

    if not candidate.is_file():
        return empty_standard_frame()

    try:

        df = pd.read_excel(candidate)

    except Exception:

        return empty_standard_frame()

    if len(df) == 0:
        return empty_standard_frame()

    df.columns = (
        df.columns
        .astype(str)
        .str.strip()
    )

    # =================================================
    # DATE
    # =================================================

    if "Date" in df.columns:

        df["Date"] = parse_excel_dates(
            df["Date"]
        )

    else:

        df["Date"] = pd.NaT

    # =================================================
    # CLIENT
    # =================================================

    c_client = first_column(
        df,
        [
            "Client Name",
            "Client name",
            "Store",
            "Customer"
        ]
    )

    if c_client:

        df["Client Name"] = (
            df[c_client]
            .astype(str)
            .str.strip()
        )

    else:

        df["Client Name"] = ""

    # =================================================
    # CROP
    # =================================================

    c_crop = first_column(
        df,
        [
            "Crop Name",
            "Crop name",
            "Crop"
        ]
    )

    if c_crop:

        df["Crop Name"] = (
            df[c_crop]
            .astype(str)
            .str.upper()
            .str.strip()
        )

    else:

        df["Crop Name"] = ""

    # =================================================
    # VARIETY
    # =================================================

    c_variety = first_column(
        df,
        [
            "Variety",
            "Colour",
            "Color"
        ]
    )

    if c_variety:

        df["Variety"] = (
            df[c_variety]
            .astype(str)
            .str.strip()
        )

    else:

        df["Variety"] = ""

    # =================================================
    # LINE
    # =================================================

    c_line = first_column(
        df,
        [
            "Lines",
            "Line"
        ]
    )

    if c_line:

        df["Line"] = (
            df[c_line]
            .astype(str)
            .str.strip()
        )

    else:

        df["Line"] = ""

    # =================================================
    # REP
    # =================================================

    c_rep = first_column(
        df,
        [
            "User",
            "Rep",
            "Sales Rep"
        ]
    )

    if c_rep:

        df["Rep"] = (
            df[c_rep]
            .astype(str)
            .str.strip()
        )

    else:

        df["Rep"] = ""

    # =================================================
    # CAVITY
    # =================================================

    c_cavity = first_column(
        df,
        [
            "Cavity",
            "Cavities"
        ]
    )

    if c_cavity:

        df["Cavity"] = pd.to_numeric(
            df[c_cavity],
            errors="coerce"
        ).fillna(0)

    else:

        df["Cavity"] = 0

    # =================================================
    # QUANTITY + AMOUNT HANDLING
    # =================================================

    filename_lower = str(path).lower()

    # ================================================
    # ORDERS FILE
    # ================================================

    if "orders" in filename_lower:

        # Amount column in orders = quantity

        c_order_qty = first_column(
            df,
            [
                "Amount",
                "Quantity",
                "Qty",
                "Units"
            ]
        )

        if c_order_qty:

            df["Quantity"] = pd.to_numeric(
                df[c_order_qty],
                errors="coerce"
            ).fillna(0)

        else:

            df["Quantity"] = 0

        # orders don't use rand sales values
        df["Amount"] = 0

        df = add_order_seed_columns(
            df
        )

    # ================================================
    # SALES + RETURNS
    # ================================================

    else:

        df = df[
            df["Date"].notna()
            & (
                df["Client Name"]
                .astype(str)
                .str.strip()
                != ""
            )
        ].copy()

        c_qty = first_column(
            df,
            [
                "Quantity",
                "Qty",
                "Units"
            ]
        )

        if c_qty:

            df["Quantity"] = pd.to_numeric(
                df[c_qty],
                errors="coerce"
            ).fillna(0)

        else:

            df["Quantity"] = 0

        c_amount = first_column(
            df,
            [
                "Amount",
                "Sales",
                "Rand",
                "Total",
                "Net Amount"
            ]
        )

        if c_amount:

            df["Amount"] = pd.to_numeric(
                df[c_amount],
                errors="coerce"
            ).fillna(0)

        else:

            df["Amount"] = 0

        if "returns" in filename_lower:

            df["Quantity"] = (
                df["Quantity"]
                .abs()
            )

            df["Amount"] = (
                df["Amount"]
                .abs()
            )

    return ensure_standard_columns(df)

# =====================================================
# LOAD DATA
# =====================================================

@st.cache_data
def load_orders(cache_version):

    return load_excel_file(
        "master_orders.xlsx"
    )


@st.cache_data
def load_sales(cache_version):

    return load_excel_file(
        "master_sales.xlsx"
    )


@st.cache_data
def load_returns(cache_version):

    return load_excel_file(
        "master_returns.xlsx"
    )


df_orders = ensure_standard_columns(
    load_orders(DATA_CACHE_VERSION)
)

df_sales = ensure_standard_columns(
    load_sales(DATA_CACHE_VERSION)
)

df_returns = ensure_standard_columns(
    load_returns(DATA_CACHE_VERSION)
)

# =====================================================
# TEXT MATCHING HELPERS
# =====================================================

def normalise_lookup_text(value):

    return re.sub(
        r"[^a-z0-9]+",
        " ",
        str(value).lower()
    ).strip()


def normalise_singular_lookup_text(value):

    text = normalise_lookup_text(value)

    words = []

    for word in text.split():

        if len(word) > 3 and word.endswith("s"):
            word = word[:-1]

        words.append(word)

    return " ".join(words)


def lookup_text_matches(needle, haystack):

    needle_norm = normalise_lookup_text(needle)
    haystack_norm = normalise_lookup_text(haystack)

    if needle_norm and needle_norm in haystack_norm:
        return True

    needle_singular = normalise_singular_lookup_text(needle)
    haystack_singular = normalise_singular_lookup_text(haystack)

    return bool(
        needle_singular
        and needle_singular in haystack_singular
    )


def line_text_matches(line, question):

    if lookup_text_matches(
        line,
        question
    ):
        return True

    line_tokens = [
        token
        for token in normalise_singular_lookup_text(line).split()
        if not re.fullmatch(
            r"\d+|\d+cm|cm",
            token
        )
    ]

    question_tokens = set(
        normalise_singular_lookup_text(question).split()
    )

    return bool(
        len(line_tokens) >= 2
        and all(
            token in question_tokens
            for token in line_tokens
        )
    )


def line_exact_matches(left, right):

    return (
        normalise_singular_lookup_text(left)
        == normalise_singular_lookup_text(right)
    )


def line_phrase_in_question(line, question):

    line_norm = normalise_lookup_text(line)
    question_norm = normalise_lookup_text(question)

    if (
        line_norm
        and re.search(
            rf"\b{re.escape(line_norm)}\b",
            question_norm
        )
    ):
        return True

    line_singular = normalise_singular_lookup_text(line)
    question_singular = normalise_singular_lookup_text(question)

    return bool(
        line_singular
        and re.search(
            rf"\b{re.escape(line_singular)}\b",
            question_singular
        )
    )


def line_group_tokens(value):

    return [
        token
        for token in normalise_singular_lookup_text(value).split()
        if not re.fullmatch(
            r"\d+|\d+cm|cm",
            token
        )
    ]


def line_dimension_tokens(value):

    return [
        token
        for token in normalise_singular_lookup_text(value).split()
        if re.fullmatch(
            r"\d+|\d+cm",
            token
        )
    ]


def line_group_matches(group_value, line_value):

    group_tokens = line_group_tokens(
        group_value
    )

    line_words = set(
        normalise_singular_lookup_text(
            line_value
        ).split()
    )

    group_dimensions = line_dimension_tokens(
        group_value
    )

    return bool(
        len(group_tokens) >= 2
        and all(
            dimension in line_words
            for dimension in group_dimensions
        )
        and all(
            token in line_words
            for token in group_tokens
        )
    )


def combined_line_question(question):

    q_norm = normalise_lookup_text(
        question
    )

    return bool(
        re.search(
            r"\b(all|both|together|combined|combine|added|add|total)\b",
            q_norm
        )
        or " and " in f" {q_norm} "
    )


def crop_family_from_question(question):

    q_norm = normalise_lookup_text(question)
    family_aliases = learned_aliases(
        "Crop Family"
    )

    for alias, family in family_aliases.items():

        if lookup_text_matches(
            alias,
            question
        ):
            return str(family).strip().upper()

    if re.search(
        r"\bpetunias\b",
        q_norm
    ):
        return "PETUNIA"

    return None


def builders_lookup_text(value):

    text = normalise_lookup_text(value)

    text = re.sub(
        r"\bbuilders\s+warehouse\b",
        "builders",
        text
    )

    text = re.sub(
        r"\bnew\b",
        "",
        text
    )

    return re.sub(
        r"\s+",
        " ",
        text
    ).strip()


def all_client_names():

    clients = pd.concat(
        [
            df_orders["Client Name"],
            df_sales["Client Name"],
            df_returns["Client Name"]
        ],
        ignore_index=True
    )

    clients = (
        clients
        .dropna()
        .astype(str)
        .str.strip()
    )

    clients = clients[
        clients != ""
    ].drop_duplicates()

    return sorted(
        clients.tolist(),
        key=lambda x: len(str(x)),
        reverse=True
    )


def learning_memory_path():

    return app_dir() / "learning_memory.json"


def load_learning_memory():

    path = learning_memory_path()

    if not path.is_file():

        return {
            "aliases": {},
            "feedback": []
        }

    try:

        with open(
            path,
            "r",
            encoding="utf-8"
        ) as f:

            memory = json.load(f)

    except Exception:

        memory = {}

    if "aliases" not in memory:

        memory["aliases"] = {}

    if "feedback" not in memory:

        memory["feedback"] = []

    if "question_intents" not in memory:

        memory["question_intents"] = {}

    if "business_rules" not in memory:

        memory["business_rules"] = {}

    if "petunia_family" not in memory["business_rules"]:

        memory["business_rules"]["petunia_family"] = {
            "crop_contains": "PETUNIA",
            "allowed_lines": PETUNIA_SEEDABLE_LINES,
            "excluded_lines": [
                "PETUNIA HYBRIDS"
            ]
        }

    return memory


def save_learning_memory(memory):

    try:

        with open(
            learning_memory_path(),
            "w",
            encoding="utf-8"
        ) as f:

            json.dump(
                memory,
                f,
                indent=2
            )

        return True

    except Exception:

        return False


def learned_aliases(column):

    memory = load_learning_memory()

    return (
        memory
        .get("aliases", {})
        .get(column, {})
    )


def learned_question_intent(question):

    memory = load_learning_memory()
    key = normalise_lookup_text(question)

    return (
        memory
        .get("question_intents", {})
        .get(key)
    )


def apply_learned_question_intent(question, intent):

    learned = learned_question_intent(
        question
    )

    if not learned:
        return intent

    updated = {
        **intent
    }

    for key, value in learned.items():

        if key in updated:
            updated[key] = value

    if (
        updated.get("line")
        and not updated.get("lines")
    ):
        updated["lines"] = [
            updated["line"]
        ]

    updated["learned"] = True

    return updated


def unique_order_values(column):

    frames = [
        df_orders,
        df_sales,
        df_returns
    ]

    values = pd.concat(
        [
            frame[column]
            for frame in frames
            if column in frame.columns
        ],
        ignore_index=True
    )

    values = values.dropna().astype(str).str.strip()

    values = values[
        values != ""
    ].drop_duplicates()

    return sorted(
        values.tolist(),
        key=lambda x: len(str(x)),
        reverse=True
    )


def detect_order_value(question, column):

    for alias, value in learned_aliases(
        column
    ).items():

        if lookup_text_matches(
            alias,
            question
        ):
            return value

    for value in unique_order_values(column):

        if lookup_text_matches(
            value,
            question
        ):
            return value

    return None


def detect_line(question):

    lines = detect_lines(
        question
    )

    if lines:
        return lines[0]

    return None


def detect_lines(question):

    combine_lines = combined_line_question(
        question
    )

    line_values = []

    for alias, value in learned_aliases(
        "Line"
    ).items():

        if lookup_text_matches(
            alias,
            question
        ):
            line_values.append(
                value
            )

    candidates = []

    for line in (
        KNOWN_LINES
        + unique_order_values("Line")
    ):

        if line not in candidates:
            candidates.append(line)

    for line in candidates:

        if line_phrase_in_question(
            line,
            question
        ):
            line_values.append(line)

    line_values = [
        line
        for index, line in enumerate(line_values)
        if line and line not in line_values[:index]
    ]

    if not line_values:
        return []

    if combine_lines:

        grouped_lines = []

        for selected_line in line_values:

            for candidate_line in candidates:

                if (
                    line_group_matches(
                        selected_line,
                        candidate_line
                    )
                    and candidate_line not in grouped_lines
                ):
                    grouped_lines.append(
                        candidate_line
                    )

        return grouped_lines or line_values

    longest_lines = []

    for line in sorted(
        line_values,
        key=lambda value: len(
            normalise_lookup_text(value)
        ),
        reverse=True
    ):

        line_norm = normalise_singular_lookup_text(
            line
        )

        if not any(
            line_norm != normalise_singular_lookup_text(existing)
            and re.search(
                rf"\b{re.escape(line_norm)}\b",
                normalise_singular_lookup_text(existing)
            )
            for existing in longest_lines
        ):
            longest_lines.append(line)

    return longest_lines


def detect_client_name(question):

    q_norm = normalise_lookup_text(question)
    q_builders_norm = builders_lookup_text(question)

    for alias, client in learned_aliases(
        "Client Name"
    ).items():

        alias_norm = normalise_lookup_text(
            alias
        )

        if alias_norm and alias_norm in q_norm:
            return client

    for client in all_client_names():

        client_norm = normalise_lookup_text(client)
        client_builders_norm = builders_lookup_text(client)

        if client_norm and client_norm in q_norm:
            return client

        if (
            client_builders_norm
            and client_builders_norm in q_builders_norm
        ):

            return client

        client_tokens = set(
            client_norm.split()
        )

        question_tokens = set(
            q_norm.split()
        )

        meaningful_tokens = [
            token
            for token in client_tokens
            if len(token) > 2
        ]

        if (
            len(meaningful_tokens) >= 2
            and all(
                token in question_tokens
                for token in meaningful_tokens
            )
        ):
            return client

    return None

# =====================================================
# INTENT DETECTION
# =====================================================

def detect_intent(question):

    ql = question.lower()

    intent = {

        "compare": False,
        "top": False,
        "seed": False,
        "metric": "quantity",
        "source": "orders",

        "line": None,
        "lines": [],
        "crop": None,
        "crop_family": None,
        "variety": None,
        "client": None,
        "rep": None,

        "year": None,
        "month": None,
        "learned": False
    }

    # =================================================
    # COMPARE
    # =================================================

    if any(
        x in ql
        for x in [
            "compare",
            "vs",
            "versus"
        ]
    ):

        intent["compare"] = True

    # =================================================
    # TOP
    # =================================================

    if any(
        x in ql
        for x in [
            "top",
            "best",
            "highest",
            "most"
        ]
    ):

        intent["top"] = True

    # =================================================
    # SOURCE
    # =================================================

    seed_order_question = bool(
        re.search(
            r"\b(seed|seeds)\b",
            ql
        )
        and not re.search(
            r"\bseedlings?\b",
            ql
        )
    )

    if seed_order_question:

        intent["seed"] = True
        intent["source"] = "orders"

    if any(
        x in ql
        for x in [
            "orders",
            "ordered",
            "order"
        ]
    ):

        intent["source"] = "orders"

    if any(
        x in ql
        for x in [
            "sales",
            "sold",
            "sell",
            "selling",
            "revenue"
        ]
    ):

        intent["source"] = "sales"

    if any(
        x in ql
        for x in [
            "returns",
            "returned",
            "refund"
        ]
    ):

        intent["source"] = "returns"

    # =================================================
    # METRIC
    # =================================================

    if any(
        x in ql
        for x in [
            "rand",
            "revenue",
            "amount",
            "value"
        ]
    ):

        intent["metric"] = "amount"

    # =================================================
    # LINE
    # =================================================

    intent["lines"] = detect_lines(
        question
    )

    if intent["lines"]:
        intent["line"] = intent["lines"][0]

    intent["crop_family"] = crop_family_from_question(
        question
    )

    # =================================================
    # CROP + VARIETY
    # =================================================

    intent["crop"] = detect_order_value(
        question,
        "Crop Name"
    )

    if intent["crop_family"]:
        intent["crop"] = None

    intent["variety"] = detect_order_value(
        question,
        "Variety"
    )

    # =================================================
    # CLIENT
    # =================================================

    intent["client"] = detect_client_name(
        question
    )

    # =================================================
    # YEAR
    # =================================================

    current_year = pd.Timestamp.today().year

    if "last year" in ql:

        intent["year"] = current_year - 1

    if "this year" in ql:

        intent["year"] = current_year

    years = re.findall(
        r"\b(20\d{2})\b",
        ql
    )

    if years:

        intent["year"] = int(years[0])

    # =================================================
    # MONTH
    # =================================================

    months = {

        "january": 1,
        "february": 2,
        "march": 3,
        "april": 4,
        "may": 5,
        "june": 6,
        "july": 7,
        "august": 8,
        "september": 9,
        "october": 10,
        "november": 11,
        "december": 12
    }

    for m, num in months.items():

        if m in ql:

            intent["month"] = num

    return apply_learned_question_intent(
        question,
        intent
    )

# =====================================================
# FILTERS
# =====================================================

def apply_filters(df, intent):

    d = ensure_standard_columns(
        df.copy()
    )

    if len(d) == 0:
        return d

    base_d = d.copy()

    has_crop_values = (
        d["Crop Name"]
        .astype(str)
        .str.strip()
        .ne("")
        .any()
    )

    has_variety_values = (
        d["Variety"]
        .astype(str)
        .str.strip()
        .ne("")
        .any()
    )

    selected_lines = intent.get("lines") or (
        [intent["line"]]
        if intent.get("line")
        else []
    )

    crop_overlaps_line = (
        selected_lines
        and intent["crop"]
        and any(
            line_group_matches(
                intent["crop"],
                line
            )
            for line in selected_lines
        )
    )

    variety_overlaps_line = (
        selected_lines
        and intent["variety"]
        and any(
            line_group_matches(
                intent["variety"],
                line
            )
            for line in selected_lines
        )
    )

    if intent["client"]:

        client_text = (
            d["Client Name"]
            .astype(str)
            .str.strip()
        )

        client_norm = client_text.map(
            normalise_lookup_text
        )

        client_builders_norm = client_text.map(
            builders_lookup_text
        )

        target_norm = normalise_lookup_text(
            intent["client"]
        )

        target_builders_norm = builders_lookup_text(
            intent["client"]
        )

        d = d[
            (client_norm == target_norm)
            | (
                client_builders_norm
                == target_builders_norm
            )
        ]

    if (
        intent.get("crop_family")
        and has_crop_values
    ):

        crop_family = str(
            intent["crop_family"]
        ).strip().upper()

        d = d[
            d["Crop Name"]
            .astype(str)
            .str.upper()
            .str.contains(
                crop_family,
                na=False
            )
        ]

        if crop_family == "PETUNIA":

            rules = (
                load_learning_memory()
                .get("business_rules", {})
                .get("petunia_family", {})
            )

            allowed_lines = rules.get(
                "allowed_lines",
                PETUNIA_SEEDABLE_LINES
            )

            excluded_lines = rules.get(
                "excluded_lines",
                [
                    "PETUNIA HYBRIDS"
                ]
            )

            d = d[
                d["Line"]
                .astype(str)
                .map(
                    lambda value: any(
                        line_exact_matches(
                            allowed_line,
                            value
                        )
                        for allowed_line in allowed_lines
                    )
                )
                .fillna(False)
                .astype(bool)
            ]

            d = d[
                ~d["Line"]
                .astype(str)
                .map(
                    lambda value: any(
                        line_exact_matches(
                            excluded_line,
                            value
                        )
                        for excluded_line in excluded_lines
                    )
                )
                .fillna(False)
                .astype(bool)
            ]

    if (
        intent["crop"]
        and has_crop_values
        and not crop_overlaps_line
    ):

        d = d[
            d["Crop Name"]
            .astype(str)
            .str.strip()
            .str.upper()
            == str(intent["crop"]).strip().upper()
        ]

    if (
        intent["variety"]
        and has_variety_values
        and not variety_overlaps_line
    ):

        d = d[
            d["Variety"]
            .astype(str)
            .str.strip()
            .str.upper()
            == str(intent["variety"]).strip().upper()
        ]

    if selected_lines:

        target_lines = [
            str(line)
            for line in selected_lines
        ]

        line_mask = (
            d["Line"]
            .astype(str)
            .map(
                lambda value: any(
                    line_exact_matches(
                        value,
                        target_line
                    )
                    for target_line in target_lines
                )
            )
            .fillna(False)
            .astype(bool)
        )

        d = d[
            line_mask
        ]

    if intent["year"]:

        d = d[
            d["Date"].dt.year
            == intent["year"]
        ]

    if intent["month"]:

        d = d[
            d["Date"].dt.month
            == intent["month"]
        ]

    if (
        len(d) == 0
        and selected_lines
        and (
            intent["crop"]
            or intent["variety"]
            or intent.get("crop_family")
        )
    ):

        fallback_intent = {
            **intent,
            "crop": None,
            "crop_family": None,
            "variety": None
        }

        return apply_filters(
            base_d,
            fallback_intent
        )

    return d


def forecast_seed_months(question):

    ql = question.lower()

    months = {

        "january": 1,
        "february": 2,
        "march": 3,
        "april": 4,
        "may": 5,
        "june": 6,
        "july": 7,
        "august": 8,
        "september": 9,
        "october": 10,
        "november": 11,
        "december": 12
    }

    today = pd.Timestamp.today()
    found = []

    for month_name, month_num in months.items():

        if month_name in ql:

            forecast_year = today.year

            if month_num < today.month:

                forecast_year = today.year + 1

            found.append(
                (
                    forecast_year,
                    month_num
                )
            )

    if found:

        return found

    horizon = 1

    if (
        "upcoming months" in ql
        or "next few months" in ql
        or "next 3 months" in ql
    ):

        horizon = 3

    targets = []

    for offset in range(
        1,
        horizon + 1
    ):

        target = today + pd.DateOffset(
            months=offset
        )

        targets.append(
            (
                target.year,
                target.month
            )
        )

    return targets


def show_seed_forecast(question, intent):

    if (
        intent.get("lines")
        and any(
            line_exact_matches(
                "PETUNIA HYBRIDS",
                line
            )
            for line in intent["lines"]
        )
    ):

        st.warning(
            "Petunia Hybrids is treated as a product line, not a seed-order crop family. Ask for 'petunias' when you want seedable Petunia crops from Seedlings, 12cm Colour Pot, and 15cm Colour Pots."
        )

        return

    target_months = forecast_seed_months(
        question
    )

    orders = apply_filters(
        df_orders,
        {
            **intent,
            "year": None,
            "month": None
        }
    )

    seed_frames = []

    for forecast_year, forecast_month in target_months:

        history_year = forecast_year - 1

        temp = orders[
            (orders["Date"].dt.year == history_year)
            & (orders["Date"].dt.month == forecast_month)
        ].copy()

        temp = temp[
            temp["Seeds Recommended"] > 0
        ].copy()

        if len(temp) == 0:

            continue

        temp["Forecast Month"] = pd.Timestamp(
            year=forecast_year,
            month=forecast_month,
            day=1
        ).strftime("%B %Y")

        temp["History Month"] = pd.Timestamp(
            year=history_year,
            month=forecast_month,
            day=1
        ).strftime("%B %Y")

        seed_frames.append(temp)

    if not seed_frames:

        period_text = ", ".join(
            [
                f"{pd.Timestamp(year=year - 1, month=month, day=1).strftime('%B %Y')}"
                for year, month in target_months
            ]
        )

        item_text = "that crop or variety"

        if intent["crop"] and intent["variety"]:

            item_text = (
                f"{intent['crop']} {intent['variety']}"
            )

        elif intent["crop"]:

            item_text = intent["crop"]

        elif intent["variety"]:

            item_text = (
                f"variety {intent['variety']}"
            )

        st.warning(
            f"We did not find any historical orders for {item_text} in {period_text}, so there is no previous-year seed demand to base this forecast on."
        )

        return

    seed_data = pd.concat(
        seed_frames,
        ignore_index=True
    )

    forecast = (
        seed_data
        .groupby(
            [
                "Forecast Month",
                "History Month",
                "Crop Name",
                "Variety",
                "Line"
            ]
        )
        .agg(
            Ordered_Units=("Quantity", "sum"),
            Seeds_Min=("Seeds Min", "sum"),
            Seeds_Max=("Seeds Max", "sum"),
            Seeds_Recommended=("Seeds Recommended", "sum")
        )
        .reset_index()
        .sort_values(
            [
                "Forecast Month",
                "Seeds_Recommended"
            ],
            ascending=[
                True,
                False
            ]
        )
    )

    forecast["Recommended_With_10pct_Growth"] = (
        forecast["Seeds_Recommended"]
        * 1.10
    ).round(0)

    totals = (
        forecast
        .groupby("Forecast Month")
        .agg(
            Seeds_Min=("Seeds_Min", "sum"),
            Seeds_Max=("Seeds_Max", "sum"),
            Seeds_Recommended=("Seeds_Recommended", "sum"),
            Recommended_With_10pct_Growth=(
                "Recommended_With_10pct_Growth",
                "sum"
            )
        )
        .reset_index()
    )

    st.subheader(
        "Seed Order Forecast"
    )

    st.caption(
        "Forecast uses the same month from the previous year. Plant to Plate is shown as a 2-3 seed range, with the recommended value using the higher number. The growth column adds 10%."
    )

    if intent["crop"]:

        st.caption(
            f"Crop: {intent['crop']}"
        )

    if intent["crop_family"]:

        st.caption(
            f"Crop Family: {intent['crop_family']}"
        )

    if intent["variety"]:

        st.caption(
            f"Variety: {intent['variety']}"
        )

    st.dataframe(
        totals,
        use_container_width=True
    )

    st.dataframe(
        forecast,
        use_container_width=True
    )

    top_seeds = (
        forecast
        .groupby("Crop Name")["Recommended_With_10pct_Growth"]
        .sum()
        .sort_values(
            ascending=False
        )
        .head(20)
    )

    fig = px.bar(
        x=top_seeds.index,
        y=top_seeds.values
    )

    st.plotly_chart(
        fig,
        use_container_width=True
    )


# =====================================================
# SOWING PLANNER EXPORT
# =====================================================

@st.cache_data(show_spinner=False)
def load_grow_weeks_table():

    path = grow_weeks_path()

    if not path.is_file():
        return pd.DataFrame(
            columns=[
                "Crop",
                "Estimated Weeks to Sale"
            ]
        )

    try:
        timing = pd.read_excel(path)
    except Exception:
        return pd.DataFrame(
            columns=[
                "Crop",
                "Estimated Weeks to Sale"
            ]
        )

    if len(timing) == 0:
        return pd.DataFrame(
            columns=[
                "Crop",
                "Estimated Weeks to Sale"
            ]
        )

    crop_col = first_column(
        timing,
        [
            "Crop",
            "Crop Name",
            "Crop / Series"
        ]
    )

    weeks_col = first_column(
        timing,
        [
            "Estimated Weeks to Sale",
            "Weeks to Sale",
            "Weeks to Tray Sale",
            "Grow Weeks"
        ]
    )

    if not crop_col or not weeks_col:
        return pd.DataFrame(
            columns=[
                "Crop",
                "Estimated Weeks to Sale"
            ]
        )

    result = timing[
        [
            crop_col,
            weeks_col
        ]
    ].copy()

    result.columns = [
        "Crop",
        "Estimated Weeks to Sale"
    ]

    result["Crop"] = (
        result["Crop"]
        .astype(str)
        .str.upper()
        .str.strip()
    )

    result["Estimated Weeks to Sale"] = pd.to_numeric(
        result["Estimated Weeks to Sale"],
        errors="coerce"
    )

    result = result[
        result["Crop"].ne("")
        & result["Estimated Weeks to Sale"].notna()
    ].copy()

    return result


def read_seed_stock_source(source):

    if source is None:

        path = seed_stock_path()

        if not path.is_file():
            return pd.DataFrame()

        name = path.name.lower()
        reader_source = path

    else:

        name = str(getattr(source, "name", "")).lower()
        reader_source = source

    try:

        if name.endswith(".csv"):
            return pd.read_csv(reader_source)

        return pd.read_excel(reader_source)

    except Exception:
        return pd.DataFrame()


def parse_at_seed_stock_layout(stock):

    if (
        "Crop" not in stock.columns
        or "Stock Take" not in stock.columns
    ):
        return None

    first_row = stock.iloc[0] if len(stock) > 0 else pd.Series(dtype="object")

    total_cols = [
        col
        for col in stock.columns
        if str(first_row.get(col, "")).strip().lower() == "total"
    ]

    qty_col = None

    if total_cols:
        qty_col = total_cols[-1]

    elif "Stock Take.1" in stock.columns:
        qty_col = "Stock Take.1"

    elif "Stock Take" in stock.columns:
        qty_col = "Stock Take"

    if qty_col is None:
        return None

    result = pd.DataFrame()

    result["Crop"] = (
        stock["Crop"]
        .astype(str)
        .str.upper()
        .str.strip()
    )

    result["Variety"] = ""

    result["Quantity On Hand"] = pd.to_numeric(
        stock[qty_col],
        errors="coerce"
    ).fillna(0)

    result = result[
        result["Crop"].ne("")
        & result["Crop"].str.upper().ne("NAN")
        & ~result["Crop"].str.lower().isin(
            [
                "price",
                "crop"
            ]
        )
        & result["Quantity On Hand"].gt(0)
    ].copy()

    result["Stock Name"] = result["Crop"]
    result["Match Text"] = result["Stock Name"].map(
        clean_timing_name
    )
    result["Family"] = result["Match Text"].map(
        seed_stock_family
    )

    return result


def load_seed_stock_table(source=None):

    stock = read_seed_stock_source(source)

    if len(stock) == 0:
        return (
            pd.DataFrame(
                columns=[
                    "Crop",
                    "Variety",
                    "Quantity On Hand",
                    "Stock Name",
                    "Match Text",
                    "Family"
                ]
            ),
            []
        )

    at_stock = parse_at_seed_stock_layout(
        stock
    )

    if at_stock is not None:
        return (
            at_stock,
            []
        )

    stock.columns = (
        stock.columns
        .astype(str)
        .str.strip()
    )

    crop_col = first_column(
        stock,
        [
            "Crop",
            "Crop Name",
            "Product",
            "Seed",
            "Seed Name",
            "Item",
            "Description"
        ]
    )

    variety_col = first_column(
        stock,
        [
            "Variety",
            "Cultivar",
            "Colour",
            "Color",
            "Series"
        ]
    )

    qty_col = first_column(
        stock,
        [
            "Quantity On Hand",
            "Qty On Hand",
            "On Hand",
            "Stock On Hand",
            "Stock",
            "Qty",
            "Quantity",
            "Seeds"
        ]
    )

    warnings = []

    if not crop_col or not qty_col:
        return (
            pd.DataFrame(
                columns=[
                    "Crop",
                    "Variety",
                    "Quantity On Hand",
                    "Stock Name",
                    "Match Text",
                    "Family"
                ]
            ),
            [
                "Seed stock list needs at least a crop/product column and a quantity/on-hand column."
            ]
        )

    result = pd.DataFrame()

    result["Crop"] = (
        stock[crop_col]
        .astype(str)
        .str.upper()
        .str.strip()
    )

    if variety_col:
        result["Variety"] = (
            stock[variety_col]
            .astype(str)
            .str.upper()
            .str.strip()
        )
    else:
        result["Variety"] = ""

    result["Quantity On Hand"] = pd.to_numeric(
        stock[qty_col],
        errors="coerce"
    ).fillna(0)

    result = result[
        result["Crop"].ne("")
        & result["Crop"].str.upper().ne("NAN")
        & result["Quantity On Hand"].gt(0)
    ].copy()

    result["Stock Name"] = result.apply(
        lambda row: product_display_name(
            row["Crop"],
            row["Variety"]
        ),
        axis=1
    )

    result["Match Text"] = result["Stock Name"].map(
        clean_timing_name
    )

    result["Family"] = result["Match Text"].map(
        seed_stock_family
    )

    return (
        result,
        warnings
    )


def seed_stock_family(value):

    words = [
        word
        for word in clean_timing_name(value).split()
        if word not in [
            "all",
            "variety",
            "serie",
            "series",
            "mix"
        ]
    ]

    if not words:
        return ""

    return words[0]


def seed_stock_match_text(crop_name):

    return clean_timing_name(
        crop_name
    )


def summarise_stock_options(options):

    if len(options) == 0:
        return ""

    sort_cols = [
        "Quantity On Hand"
    ]

    sort_ascending = [
        False
    ]

    if "Match Score" in options.columns:
        sort_cols = [
            "Match Score",
            "Quantity On Hand"
        ]
        sort_ascending = [
            False,
            False
        ]

    return ", ".join(
        options
        .sort_values(
            sort_cols,
            ascending=sort_ascending
        )
        .head(4)
        .apply(
            lambda row: f"{row['Stock Name']} ({row['Quantity On Hand']:,.0f})",
            axis=1
        )
        .tolist()
    )


def seed_stock_match_score(requested_text, stock_text):

    requested_words = set(
        clean_timing_name(
            requested_text
        ).split()
    )

    stock_words = set(
        clean_timing_name(
            stock_text
        ).split()
    )

    if not requested_words or not stock_words:
        return 0

    return len(
        requested_words.intersection(
            stock_words
        )
    )


def apply_seed_stock_to_plan(plan, stock, germ_pct, growth_pct):

    if len(plan) == 0:
        return (
            plan,
            []
        )

    plan = plan.copy()

    plan["Estimated Seeds Needed"] = (
        (
            plan["Base Forecast Output Units"].astype(float)
            * plan["Output Divisor"].astype(float)
            * (1 + float(growth_pct))
        )
        / max(
            float(germ_pct),
            0.01
        )
    ).round().astype(int)

    plan["Seed Stock Status"] = "No stock list"
    plan["Seed Stock Match"] = ""
    plan["Seed Stock On Hand"] = 0

    if len(stock) == 0:
        return (
            plan,
            []
        )

    available = stock.copy()
    warnings = []

    for index, row in plan.iterrows():

        needed = int(
            row["Estimated Seeds Needed"]
        )

        requested_text = seed_stock_match_text(
            row["Crop Name"]
        )

        requested_family = seed_stock_family(
            row["Crop Name"]
        )

        exact = available[
            available["Match Text"].eq(requested_text)
        ].copy()

        if len(exact) > 0:
            exact["Match Score"] = 999

        match_type = "Exact"
        options = exact

        if len(options) == 0:

            options = available[
                available["Family"].eq(requested_family)
            ].copy()

            options["Match Score"] = options["Match Text"].map(
                lambda stock_text: seed_stock_match_score(
                    requested_text,
                    stock_text
                )
            )

            match_type = "Close variety"

        available_qty = int(
            options["Quantity On Hand"].sum()
        ) if len(options) > 0 else 0

        match_summary = summarise_stock_options(
            options
        )

        if available_qty <= 0:
            status = "Missing"
            warnings.append(
                f"No seed stock found for {row['Crop Name']}."
            )

        elif available_qty < needed:
            status = "Low stock"
            warnings.append(
                f"Low seed stock for {row['Crop Name']}: need {needed:,.0f}, stock match has {available_qty:,.0f}."
            )

        elif match_type == "Close variety":
            status = "Use close variety"
            warnings.append(
                f"{row['Crop Name']} has no exact seed stock match; close stock can cover it: {match_summary}."
            )

        else:
            status = "In stock"

        plan.loc[index, "Seed Stock Status"] = status
        plan.loc[index, "Seed Stock Match"] = match_summary
        plan.loc[index, "Seed Stock On Hand"] = available_qty

        if status not in [
            "Missing",
            "Low stock"
        ]:

            remaining = needed

            for stock_index in options.sort_values(
                [
                    "Match Score",
                    "Quantity On Hand"
                ],
                ascending=[
                    False,
                    False
                ]
            ).index:

                take = min(
                    remaining,
                    int(available.loc[stock_index, "Quantity On Hand"])
                )

                available.loc[stock_index, "Quantity On Hand"] -= take
                remaining -= take

                if remaining <= 0:
                    break

    plan["Notes"] = (
        plan["Notes"].astype(str)
        + " Seed stock: "
        + plan["Seed Stock Status"].astype(str)
        + ". "
        + plan["Seed Stock Match"].astype(str)
    )

    return (
        plan,
        warnings[:20]
    )


def build_timing_lookup(timing):

    rows = []
    exact = {}

    for _, row in timing.iterrows():

        timing_name = str(
            row["Crop"]
        )

        weeks = int(
            round(
                float(row["Estimated Weeks to Sale"])
            )
        )

        timing_norm = clean_timing_name(
            timing_name
        )

        rows.append(
            (
                timing_name,
                weeks,
                timing_norm
            )
        )

        if timing_norm and timing_norm not in exact:
            exact[timing_norm] = (
                weeks,
                timing_name
            )

    return {
        "rows": rows,
        "exact": exact
    }


def clean_timing_name(value):

    text = normalise_singular_lookup_text(
        value
    )

    text = re.sub(
        r"\ball variet(?:y|ie)\b",
        "",
        text
    )

    text = re.sub(
        r"\ball series\b",
        "",
        text
    )

    return re.sub(
        r"\s+",
        " ",
        text
    ).strip()


def lookup_grow_weeks(crop_name, variety, timing):

    result = lookup_grow_timing(
        crop_name,
        variety,
        timing
    )

    if result:
        return result[0]

    return None


def lookup_grow_timing(crop_name, variety, timing, timing_lookup=None):

    if len(timing) == 0:
        return None

    crop_text = str(crop_name).strip().upper()
    variety_text = str(variety).strip().upper()

    search_values = [
        f"{crop_text} {variety_text}".strip(),
        crop_text
    ]

    if timing_lookup is None:
        timing_lookup = build_timing_lookup(
            timing
        )

    timing_rows = timing_lookup["rows"]
    exact_lookup = timing_lookup["exact"]

    for search_value in search_values:

        search_norm = clean_timing_name(
            search_value
        )

        if search_norm in exact_lookup:
            return exact_lookup[search_norm]

    for search_value in search_values:

        search_norm = clean_timing_name(
            search_value
        )

        for timing_name, weeks, timing_norm in timing_rows:

            if (
                timing_norm
                and (
                    search_norm.startswith(
                        timing_norm
                    )
                    or timing_norm.startswith(
                        search_norm
                    )
                )
            ):
                return (
                    int(weeks),
                    timing_name
                )

    first_word = clean_timing_name(
        crop_text
    ).split()

    if first_word:

        first_word = first_word[0]

        for timing_name, weeks, timing_norm in timing_rows:

            if timing_norm == first_word:
                return (
                    int(weeks),
                    timing_name
                )

    return None


def iso_week_start(year, week):

    target_year = int(year)
    target_week = int(week)

    while True:

        try:
            return date.fromisocalendar(
                target_year,
                target_week,
                1
            )

        except ValueError:

            last_week = date(
                target_year,
                12,
                28
            ).isocalendar().week

            if target_week > last_week:
                target_week -= last_week
                target_year += 1
            else:
                target_week = 1


def output_divisor_for_order(line, cavity):

    line_text = str(line).upper()

    if "PLANT TO PLATE" in line_text:
        return None

    if "PETUNIA HYBRIDS" in line_text:
        return None

    if "SEEDLINGS" in line_text:

        if float(cavity or 0) == 4:
            return 4

        return 6

    if "12CM" in line_text:
        return 1

    if "15CM" in line_text:
        return 2

    return None


def add_output_divisor_column(orders):

    orders = orders.copy()

    line = (
        orders["Line"]
        .astype(str)
        .str.upper()
    )

    cavity = pd.to_numeric(
        orders["Cavity"],
        errors="coerce"
    ).fillna(0)

    output_divisor = pd.Series(
        pd.NA,
        index=orders.index,
        dtype="Float64"
    )

    seedling = (
        line.str.contains(
            "SEEDLINGS",
            na=False
        )
        & ~line.str.contains(
            "PETUNIA HYBRIDS",
            na=False
        )
    )

    output_divisor.loc[seedling] = 6
    output_divisor.loc[seedling & cavity.eq(4)] = 4

    output_divisor.loc[
        line.str.contains(
            "12CM",
            na=False
        )
    ] = 1

    output_divisor.loc[
        line.str.contains(
            "15CM",
            na=False
        )
    ] = 2

    output_divisor.loc[
        line.str.contains(
            "PLANT TO PLATE",
            na=False
        )
    ] = pd.NA

    output_divisor.loc[
        line.str.contains(
            "PETUNIA HYBRIDS",
            na=False
        )
    ] = pd.NA

    orders["Output Divisor"] = output_divisor

    return orders


def product_display_name(crop_name, variety):

    crop = str(crop_name).strip().upper()
    var = str(variety).strip().upper()

    if not var or var in [
        "NAN",
        "NONE"
    ]:
        return crop

    if var in crop:
        return crop

    return f"{crop} {var}"


def forecast_weekly_sowing_rows(
    current_week,
    plan_year,
    planning_tray_cells=595,
    germ_pct=0.9,
    min_tray_fraction=0.5,
    max_rows=50
):

    timing = load_grow_weeks_table()
    timing_lookup = build_timing_lookup(
        timing
    )
    warnings = []

    if len(timing) == 0:
        warnings.append(
            f"Missing or empty {GROW_WEEKS_FILE}; using 6 weeks for all planner rows."
        )

    orders = ensure_standard_columns(
        df_orders.copy()
    )

    orders = orders[
        orders["Date"].notna()
        & orders["Quantity"].gt(0)
    ].copy()

    if len(orders) == 0:
        return (
            pd.DataFrame(),
            [
                "No order data was available for the sowing forecast."
            ]
        )

    orders = add_output_divisor_column(
        orders
    )

    orders = orders[
        orders["Output Divisor"].notna()
    ].copy()

    orders["Planner Crop"] = orders.apply(
        lambda row: product_display_name(
            row["Crop Name"],
            row["Variety"]
        ),
        axis=1
    )

    timing_keys = (
        orders[
            [
                "Crop Name",
                "Variety"
            ]
        ]
        .drop_duplicates()
    )

    timing_match_map = {}

    for _, key_row in timing_keys.iterrows():

        key = (
            str(key_row["Crop Name"]),
            str(key_row["Variety"])
        )

        timing_match_map[key] = lookup_grow_timing(
            key[0],
            key[1],
            timing,
            timing_lookup
        )

    timing_matches = pd.Series(
        [
            timing_match_map.get(
                (
                    str(crop_name),
                    str(variety)
                )
            )
            for crop_name, variety in zip(
                orders["Crop Name"],
                orders["Variety"]
            )
        ],
        index=orders.index
    )

    missing_timing = orders[
        timing_matches.isna()
    ]

    if len(missing_timing) > 0 and len(timing) > 0:

        for crop_name in (
            missing_timing["Crop Name"]
            .dropna()
            .astype(str)
            .str.upper()
            .drop_duplicates()
            .head(20)
        ):
            warnings.append(
                f"No grow weeks found for {crop_name}; using 6 weeks."
            )

    orders["Grow Weeks"] = timing_matches.map(
        lambda value: value[0] if value else 6
    )

    orders["Timing Crop"] = timing_matches.map(
        lambda value: value[1] if value else "Default 6 weeks"
    )

    orders = orders[
        orders["Planner Crop"].notna()
        & orders["Planner Crop"].astype(str).str.strip().ne("")
    ].copy()

    if len(orders) == 0:
        return (
            pd.DataFrame(),
            warnings
            + [
                "No orders were available after preparing planner crop names."
            ]
        )

    orders["Year"] = orders["Date"].dt.year
    orders["Month"] = orders["Date"].dt.month

    monthly = (
        orders
        .groupby(
            [
                "Year",
                "Month",
                "Line",
                "Planner Crop",
                "Timing Crop",
                "Grow Weeks",
                "Output Divisor"
            ],
            dropna=False
        )["Quantity"]
        .sum()
        .reset_index()
    )

    varieties = (
        orders
        .groupby(
            [
                "Year",
                "Month",
                "Line",
                "Planner Crop",
                "Timing Crop",
                "Grow Weeks",
                "Output Divisor"
            ],
            dropna=False
        )["Crop Name"]
        .apply(
            lambda values: ", ".join(
                values
                .dropna()
                .astype(str)
                .str.upper()
                .drop_duplicates()
                .head(5)
            )
        )
        .reset_index(name="Included Crops")
    )

    monthly = monthly.merge(
        varieties,
        on=[
            "Year",
            "Month",
            "Line",
            "Planner Crop",
            "Timing Crop",
            "Grow Weeks",
            "Output Divisor"
        ],
        how="left"
    )

    rows = []

    product_groups = (
        orders[
            [
                "Line",
                "Planner Crop",
                "Timing Crop",
                "Grow Weeks",
                "Output Divisor"
            ]
        ]
        .drop_duplicates()
    )

    for _, product in product_groups.iterrows():

        grow_weeks = int(
            product["Grow Weeks"]
        )

        ready_week_start = iso_week_start(
            plan_year,
            int(current_week) + grow_weeks
        )

        target_year = ready_week_start.year
        target_month = ready_week_start.month
        history_year = target_year - 1

        demand = monthly[
            (monthly["Year"] == history_year)
            & (monthly["Month"] == target_month)
            & (monthly["Line"].astype(str) == str(product["Line"]))
            & (monthly["Planner Crop"].astype(str) == str(product["Planner Crop"]))
            & (monthly["Timing Crop"].astype(str) == str(product["Timing Crop"]))
            & (monthly["Grow Weeks"].astype(float) == float(product["Grow Weeks"]))
            & (
                monthly["Output Divisor"].astype(float)
                == float(product["Output Divisor"])
            )
        ]

        monthly_qty = demand["Quantity"].sum()

        if monthly_qty <= 0:
            continue

        output_divisor = int(
            product["Output Divisor"]
        )

        practical_batch = int(
            max(
                1,
                round(
                    (
                        float(planning_tray_cells)
                        * float(germ_pct)
                        / output_divisor
                    )
                    * float(min_tray_fraction)
                )
            )
        )

        weekly_qty = int(
            max(
                1,
                round(
                    monthly_qty / 4
                )
            )
        )

        if monthly_qty < practical_batch:
            continue

        if weekly_qty < practical_batch:
            weekly_qty = practical_batch

        included_crops = ""
        timing_source = ""

        if len(demand) > 0:
            included_crops = str(
                demand["Included Crops"].iloc[0]
            )

        timing_source = str(
            product.get(
                "Timing Crop",
                ""
            )
        )

        rows.append({
            "Line": product["Line"],
            "Crop Name": str(
                product["Planner Crop"]
            ),
            "Grow Weeks": grow_weeks,
            "Week Ready For": int(current_week) + grow_weeks,
            "Grow Tray": "",
            "Base Forecast Output Units": weekly_qty,
            "Output Divisor": output_divisor,
            "History Month": pd.Timestamp(
                year=history_year,
                month=target_month,
                day=1
            ).strftime("%B %Y"),
            "Monthly Demand": monthly_qty,
            "Demand Priority Score": monthly_qty,
            "Notes": (
                f"Orders only. {pd.Timestamp(year=history_year, month=target_month, day=1).strftime('%B %Y')} "
                f"demand {monthly_qty:,.0f} split across 4 weeks. "
                f"Minimum practical batch {practical_batch:,.0f} output units. "
                f"Grow timing: {timing_source}. "
                f"Includes: {included_crops}"
            )
        })

    plan = pd.DataFrame(rows)

    if len(plan) == 0:
        return (
            plan,
            warnings[:20]
            + [
                "No sowing rows were generated after applying grow weeks, line exclusions, and order demand."
            ]
        )

    plan = plan.sort_values(
        [
            "Demand Priority Score",
            "Monthly Demand",
            "Base Forecast Output Units",
            "Week Ready For",
            "Crop Name"
        ],
        ascending=[
            False,
            False,
            False,
            True,
            True
        ]
    ).reset_index(drop=True)

    plan["Sow Priority"] = range(
        1,
        len(plan) + 1
    )

    if max_rows and len(plan) > int(max_rows):

        warnings.append(
            f"Planner found {len(plan)} rows. Export is limited to the top {int(max_rows)} by priority."
        )

        plan = plan.head(
            int(max_rows)
        ).copy()

    return (
        plan,
        warnings[:20]
    )


def copy_cell_template(source_cell, target_cell):

    if source_cell.has_style:
        target_cell._style = copy(
            source_cell._style
        )

    target_cell.number_format = source_cell.number_format
    target_cell.font = copy(source_cell.font)
    target_cell.fill = copy(source_cell.fill)
    target_cell.border = copy(source_cell.border)
    target_cell.alignment = copy(source_cell.alignment)

    if (
        isinstance(source_cell.value, str)
        and source_cell.value.startswith("=")
    ):
        target_cell.value = Translator(
            source_cell.value,
            origin=source_cell.coordinate
        ).translate_formula(
            target_cell.coordinate
        )
    else:
        target_cell.value = None


def ensure_planner_rows(wb, required_rows):

    input_start = 9
    input_template_row = 58
    weekly_start = 4
    weekly_template_row = 53

    if required_rows <= 50:
        return

    ws_inputs = wb["Inputs"]
    ws_weekly = wb["Weekly Plan"]

    for offset in range(
        50,
        required_rows
    ):

        input_row = input_start + offset
        weekly_row = weekly_start + offset

        for col in range(
            1,
            ws_inputs.max_column + 1
        ):
            copy_cell_template(
                ws_inputs.cell(
                    input_template_row,
                    col
                ),
                ws_inputs.cell(
                    input_row,
                    col
                )
            )

        for col in range(
            1,
            16
        ):
            copy_cell_template(
                ws_weekly.cell(
                    weekly_template_row,
                    col
                ),
                ws_weekly.cell(
                    weekly_row,
                    col
                )
            )

    last_weekly_row = weekly_start + required_rows - 1

    ws_weekly["R5"] = f"=SUM(L4:L{last_weekly_row})"
    ws_weekly["R6"] = f"=SUMIF(G4:G{last_weekly_row},595,M4:M{last_weekly_row})"
    ws_weekly["R7"] = f"=SUMIF(G4:G{last_weekly_row},512,M4:M{last_weekly_row})"
    ws_weekly["R8"] = f"=SUMIF(G4:G{last_weekly_row},128,M4:M{last_weekly_row})"
    ws_weekly["R9"] = f"=SUM(N4:N{last_weekly_row})"


def sow_list_seed_quantity(row):

    seeds_needed = int(
        row.get(
            "Estimated Seeds Needed",
            0
        ) or 0
    )

    if seeds_needed <= 0:
        seeds_needed = int(
            row.get(
                "Base Forecast Output Units",
                0
            ) or 0
        ) * int(
            row.get(
                "Output Divisor",
                1
            ) or 1
        )

    return max(
        1,
        seeds_needed
    )


def sow_list_quantity_formula(excel_row, seeds_needed):

    return (
        f'=IF(G{excel_row}="","",ROUNDUP({int(seeds_needed)}/G{excel_row},0))'
    )


def style_weekly_sow_list_sheet(ws):

    header_fill = PatternFill(
        "solid",
        fgColor="A6A6A6"
    )
    list_fill = PatternFill(
        "solid",
        fgColor="D9D9D9"
    )
    white_fill = PatternFill(
        "solid",
        fgColor="FFFFFF"
    )
    thin = Side(
        style="thin",
        color="000000"
    )
    border = Border(
        left=thin,
        right=thin,
        top=thin,
        bottom=thin
    )

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(
            bold=False
        )
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )
        cell.border = border

    for row in ws.iter_rows(
        min_row=2,
        max_row=ws.max_row
    ):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(
                vertical="center",
                wrap_text=True
            )

        row[0].fill = header_fill
        row[1].fill = list_fill
        row[7].fill = white_fill

        for col_index in [
            1,
            2,
            4,
            7,
            8,
            9,
            11,
            12,
            14
        ]:
            row[col_index - 1].alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True
            )

    widths = {
        "A": 8,
        "B": 10,
        "C": 12,
        "D": 8,
        "E": 34,
        "F": 16,
        "G": 8,
        "H": 12,
        "I": 12,
        "J": 26,
        "K": 10,
        "L": 14,
        "M": 16,
        "N": 14
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_margins = PageMargins(
        left=0.25,
        right=0.25,
        top=0.5,
        bottom=0.5,
        header=0.2,
        footer=0.2
    )
    ws.print_title_rows = "1:1"


def add_weekly_sow_list_sheet(
    wb,
    plan,
    current_week,
    plan_year,
    planning_tray_cells
):

    sheet_name = "Weekly Sow List"

    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    ws = wb.create_sheet(
        sheet_name
    )

    headers = [
        "Year",
        "Sowing List #",
        "Date Sowed",
        "Priority",
        "Full Name Of Crop & Colour",
        "Production Line",
        "Cavity ",
        "Total Crops Sowed",
        "Quantity Needed",
        "Comments",
        "Batch #",
        "Date out Germ Room",
        "Person out of Germ Room",
        "Seed Stock"
    ]

    ws.append(
        headers
    )

    sow_list = plan.sort_values(
        [
            "Sow Priority",
            "Line",
            "Crop Name"
        ],
        ascending=[
            True,
            True,
            True
        ]
    )

    for _, row in sow_list.iterrows():

        input_row = 9 + int(
            row.name
        )

        seeds_needed = sow_list_seed_quantity(
            row
        )
        excel_row = ws.max_row + 1

        comments = (
            f"Ready wk {int(row.get('Week Ready For', 0) or 0)}. "
            f"Grow {int(row.get('Grow Weeks', 0) or 0)} wks. "
            f"Demand {float(row.get('Monthly Demand', 0) or 0):,.0f}."
        ).strip()

        ws.append([
            int(plan_year),
            int(current_week),
            None,
            int(row["Sow Priority"]),
            str(row["Crop Name"]).upper(),
            row.get("Line", ""),
            f"='Inputs'!G{input_row}",
            None,
            sow_list_quantity_formula(
                excel_row,
                seeds_needed
            ),
            comments,
            None,
            None,
            None,
            row.get("Seed Stock Status", "")
        ])

    style_weekly_sow_list_sheet(
        ws
    )


def build_sowing_planner_workbook(
    plan,
    current_week,
    germ_pct,
    growth_pct,
    plan_year,
    planning_tray_cells
):

    path = template_path()

    if not path.is_file():
        return None

    wb = load_workbook(
        path
    )

    add_weekly_sow_list_sheet(
        wb,
        plan,
        current_week,
        plan_year,
        planning_tray_cells
    )

    ensure_planner_rows(
        wb,
        len(plan)
    )

    ws = wb["Inputs"]

    ws["B2"] = int(current_week)
    ws["B3"] = float(germ_pct)
    ws["B4"] = float(growth_pct)

    manual_cols = [
        "A",
        "B",
        "C",
        "E",
        "G",
        "H",
        "K",
        "U"
    ]

    final_row = 8 + max(
        50,
        len(plan)
    )

    for row in range(
        9,
        final_row + 1
    ):

        for col in manual_cols:
            ws[f"{col}{row}"] = None

    for index, row in plan.iterrows():

        excel_row = 9 + index

        ws[f"A{excel_row}"] = index + 1
        ws[f"B{excel_row}"] = row["Crop Name"]
        ws[f"C{excel_row}"] = int(row["Sow Priority"])
        ws[f"E{excel_row}"] = int(row["Grow Weeks"])
        ws[f"G{excel_row}"] = None
        ws[f"H{excel_row}"] = int(row["Base Forecast Output Units"])
        ws[f"K{excel_row}"] = int(row["Output Divisor"])
        ws[f"U{excel_row}"] = row["Notes"]

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output


def render_sowing_planner_export():

    st.header(
        "Weekly Sowing Planner Export"
    )

    st.caption(
        "Uses orders, grow weeks, and optional seed stock. The order date is treated as the ready-to-sell period, then demand is averaged across the month to create this week's sowing quantities."
    )

    template_available = template_path().is_file()
    timing_available = grow_weeks_path().is_file()

    if not template_available:
        st.warning(
            f"Missing {SOWING_PLANNER_TEMPLATE}. Add it to the app folder or Documents folder."
        )

    if not timing_available:
        st.warning(
            f"Missing {GROW_WEEKS_FILE}. Add the crop grow-week lookup to the app folder, Documents, or Downloads."
        )

    if seed_stock_path().is_file():
        st.caption(
            f"Using local seed stock file: {seed_stock_path().name}"
        )

    today = pd.Timestamp.today()
    default_week = int(
        today.isocalendar().week
    )

    col1, col2, col3 = st.columns(3)

    current_week = col1.number_input(
        "Current production week",
        min_value=1,
        max_value=53,
        value=default_week,
        step=1
    )

    plan_year = col2.number_input(
        "Planning year",
        min_value=2020,
        max_value=2100,
        value=int(today.year),
        step=1
    )

    germ_pct = col3.number_input(
        "Default germination %",
        min_value=0.1,
        max_value=1.0,
        value=0.9,
        step=0.01
    )

    growth_pct = st.number_input(
        "Sales growth %",
        min_value=-1.0,
        max_value=5.0,
        value=0.05,
        step=0.01
    )

    col4, col5, col6 = st.columns(3)

    planning_tray_cells = col4.selectbox(
        "Planning tray size for minimum batches",
        [
            595,
            512
        ],
        index=0
    )

    min_tray_fraction = col5.slider(
        "Smallest practical batch",
        min_value=0.25,
        max_value=1.0,
        value=0.5,
        step=0.25
    )

    max_rows = col6.number_input(
        "Maximum planner rows",
        min_value=10,
        max_value=200,
        value=50,
        step=10
    )

    if st.button(
        "Build sowing planner"
    ):

        plan, warnings = forecast_weekly_sowing_rows(
            int(current_week),
            int(plan_year),
            int(planning_tray_cells),
            float(germ_pct),
            float(min_tray_fraction),
            int(max_rows)
        )

        stock, stock_warnings = load_seed_stock_table()

        plan, seed_stock_warnings = apply_seed_stock_to_plan(
            plan,
            stock,
            float(germ_pct),
            float(growth_pct)
        )

        for warning in (
            warnings
            + stock_warnings
            + seed_stock_warnings
        ):
            st.warning(
                warning
            )

        if len(plan) == 0:
            return

        st.subheader(
            "Planner Preview"
        )

        if len(stock) > 0:

            status_counts = (
                plan["Seed Stock Status"]
                .value_counts()
                .to_dict()
            )

            stock_col1, stock_col2, stock_col3 = st.columns(3)

            stock_col1.metric(
                "Seed rows in stock",
                f"{status_counts.get('In stock', 0):,.0f}"
            )

            stock_col2.metric(
                "Close variety rows",
                f"{status_counts.get('Use close variety', 0):,.0f}"
            )

            stock_col3.metric(
                "Missing / low rows",
                f"{status_counts.get('Missing', 0) + status_counts.get('Low stock', 0):,.0f}"
            )

        preview_cols = [
            "Sow Priority",
            "Line",
            "Crop Name",
            "Demand Priority Score",
            "Grow Weeks",
            "Week Ready For",
            "Base Forecast Output Units",
            "Output Divisor",
            "Estimated Seeds Needed",
            "Seed Stock Status",
            "Seed Stock Match",
            "Seed Stock On Hand",
            "History Month",
            "Monthly Demand"
        ]

        st.dataframe(
            plan[preview_cols],
            use_container_width=True
        )

        workbook = build_sowing_planner_workbook(
            plan,
            int(current_week),
            float(germ_pct),
            float(growth_pct),
            int(plan_year),
            int(planning_tray_cells)
        )

        if workbook is None:
            st.warning(
                "Could not build the workbook because the template was not found."
            )
            return

        filename = (
            f"AT_Nursery_Sowing_Planner_Week_{int(current_week)}_"
            f"{int(plan_year)}.xlsx"
        )

        st.download_button(
            "Download filled sowing planner",
            data=workbook,
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )


def save_alias(memory, column, alias, value):

    alias = str(alias).strip()
    value = str(value).strip()

    if not alias or not value:

        return

    if column not in memory["aliases"]:

        memory["aliases"][column] = {}

    memory["aliases"][column][alias] = value


def save_question_intent(memory, question, intent_update):

    key = normalise_lookup_text(
        question
    )

    if not key:
        return

    cleaned = {}

    for intent_key, value in intent_update.items():

        if value not in [
            "",
            None
        ]:
            cleaned[intent_key] = value

    if not cleaned:
        return

    memory["question_intents"][key] = cleaned


def intent_summary_rows(intent):

    labels = {
        "source": "Dataset",
        "client": "Client",
        "lines": "Lines",
        "line": "Line",
        "crop": "Crop",
        "crop_family": "Crop Family",
        "variety": "Variety",
        "year": "Year",
        "month": "Month",
        "metric": "Metric"
    }

    rows = []

    for key, label in labels.items():

        value = intent.get(key)

        if (
            key == "line"
            and intent.get("lines")
        ):
            continue

        if isinstance(
            value,
            list
        ):
            value = ", ".join(
                [
                    str(item)
                    for item in value
                ]
            )

        if value not in [
            None,
            ""
        ]:
            rows.append({
                "Filter": label,
                "Value": value
            })

    return pd.DataFrame(
        rows
    )


def month_label(month):

    if not month:
        return None

    try:
        return pd.Timestamp(
            year=2000,
            month=int(month),
            day=1
        ).strftime("%B")

    except Exception:
        return str(month)


def stable_question_for_exact_learning(question):

    q_norm = normalise_lookup_text(
        question
    )

    moving_date_words = [
        "this year",
        "last year",
        "next year",
        "previous year",
        "this month",
        "last month",
        "next month",
        "today",
        "yesterday",
        "tomorrow",
        "current"
    ]

    return not any(
        phrase in q_norm
        for phrase in moving_date_words
    )


def intent_for_memory(intent):

    return {
        "source": intent.get("source"),
        "line": intent.get("line"),
        "lines": intent.get("lines"),
        "crop": intent.get("crop"),
        "crop_family": intent.get("crop_family"),
        "variety": intent.get("variety"),
        "client": intent.get("client"),
        "year": intent.get("year"),
        "month": intent.get("month")
    }


def save_quick_feedback(question, intent, rating, notes):

    memory = load_learning_memory()

    memory["feedback"].append({
        "question": question,
        "useful": rating,
        "notes": notes,
        "quick_feedback": True,
        "detected": intent_for_memory(
            intent
        )
    })

    if (
        rating == "Correct"
        and stable_question_for_exact_learning(
            question
        )
    ):

        save_question_intent(
            memory,
            question,
            intent_for_memory(
                intent
            )
        )

    return save_learning_memory(
        memory
    )


def render_understanding_panel(question, intent):

    with st.expander(
        "What I understood",
        expanded=True
    ):

        dataset_label = str(
            intent.get("source", "orders")
        ).title()

        st.write(
            f"I will answer this from **{dataset_label}**."
        )

        if intent.get("client"):

            st.write(
                f"Client: **{intent['client']}**"
            )

        if intent.get("lines"):

            st.write(
                f"Line: **{', '.join(intent['lines'])}**"
            )

        elif intent.get("line"):

            st.write(
                f"Line: **{intent['line']}**"
            )

        if intent.get("crop_family"):

            st.write(
                f"Crop family: **{intent['crop_family']}**"
            )

        if intent.get("crop"):

            st.write(
                f"Crop: **{intent['crop']}**"
            )

        if intent.get("variety"):

            st.write(
                f"Variety: **{intent['variety']}**"
            )

        period_parts = []

        if intent.get("month"):

            period_parts.append(
                month_label(
                    intent["month"]
                )
            )

        if intent.get("year"):

            period_parts.append(
                str(intent["year"])
            )

        if period_parts:

            st.write(
                f"Period: **{' '.join(period_parts)}**"
            )

        st.caption(
            "If this understanding is wrong, mark it below or use the detailed learning section after the answer."
        )

        if intent.get("learned"):

            st.success(
                "I used a saved correction for this question."
            )

        with st.form(
            "quick_understanding_feedback_form"
        ):

            rating = st.radio(
                "Was this understanding and answer correct?",
                [
                    "Correct",
                    "Partly",
                    "Wrong"
                ],
                horizontal=True
            )

            notes = st.text_input(
                "Optional note"
            )

            submitted = st.form_submit_button(
                "Save quick feedback"
            )

            if submitted:

                if save_quick_feedback(
                    question,
                    intent,
                    rating,
                    notes
                ):

                    st.success(
                        "Saved. I will use this feedback to improve future questions."
                    )

                else:

                    st.warning(
                        "I could not save the feedback file."
                    )


def render_learning_feedback(question, intent):

    with st.expander(
        "Help the app learn from this question"
    ):

        st.caption(
            "Save corrections here when the app misunderstands wording. Exact question corrections are applied before future answers."
        )

        with st.form(
            "learning_feedback_form"
        ):

            useful = st.radio(
                "Was this answer correct or useful?",
                [
                    "Yes",
                    "No",
                    "Partly"
                ],
                horizontal=True
            )

            notes = st.text_area(
                "Correction notes"
            )

            st.caption(
                "Aliases teach reusable wording. Exact question corrections teach this full question."
            )

            source_value = st.selectbox(
                "Correct dataset",
                [
                    "",
                    "orders",
                    "sales",
                    "returns"
                ],
                index=0
            )

            line_alias = st.text_input(
                "Line alias phrase"
            )

            line_options = [
                ""
            ] + unique_order_values(
                "Line"
            )

            line_value = st.selectbox(
                "Correct line",
                line_options,
                index=0
            )

            crop_alias = st.text_input(
                "Crop alias phrase"
            )

            crop_options = [
                ""
            ] + unique_order_values(
                "Crop Name"
            )

            crop_value = st.selectbox(
                "Correct crop name",
                crop_options,
                index=0
            )

            crop_family_alias = st.text_input(
                "Crop family alias phrase"
            )

            crop_family_value = st.selectbox(
                "Correct crop family",
                [
                    "",
                    "PETUNIA"
                ],
                index=0
            )

            variety_alias = st.text_input(
                "Variety alias phrase"
            )

            variety_options = [
                ""
            ] + unique_order_values(
                "Variety"
            )

            variety_value = st.selectbox(
                "Correct variety",
                variety_options,
                index=0
            )

            client_alias = st.text_input(
                "Client alias phrase"
            )

            client_options = [
                ""
            ] + all_client_names()

            client_value = st.selectbox(
                "Correct client name",
                client_options,
                index=0
            )

            year_value = st.number_input(
                "Correct year",
                min_value=0,
                max_value=2100,
                value=0,
                step=1
            )

            month_value = st.selectbox(
                "Correct month",
                [
                    "",
                    "1",
                    "2",
                    "3",
                    "4",
                    "5",
                    "6",
                    "7",
                    "8",
                    "9",
                    "10",
                    "11",
                    "12"
                ],
                index=0
            )

            remember_exact = st.checkbox(
                "Remember these corrections for this exact question",
                value=True
            )

            submitted = st.form_submit_button(
                "Save feedback"
            )

            if submitted:

                memory = load_learning_memory()

                memory["feedback"].append({
                    "question": question,
                    "useful": useful,
                    "notes": notes,
                    "detected": {
                        "source": intent["source"],
                        "line": intent["line"],
                        "lines": intent.get("lines"),
                        "crop": intent["crop"],
                        "crop_family": intent.get("crop_family"),
                        "variety": intent["variety"],
                        "client": intent["client"],
                        "year": intent["year"],
                        "month": intent["month"]
                    }
                })

                save_alias(
                    memory,
                    "Line",
                    line_alias,
                    line_value
                )

                save_alias(
                    memory,
                    "Crop Name",
                    crop_alias,
                    crop_value
                )

                save_alias(
                    memory,
                    "Crop Family",
                    crop_family_alias,
                    crop_family_value
                )

                save_alias(
                    memory,
                    "Variety",
                    variety_alias,
                    variety_value
                )

                save_alias(
                    memory,
                    "Client Name",
                    client_alias,
                    client_value
                )

                if remember_exact:

                    exact_intent = {
                        "source": source_value,
                        "line": line_value,
                        "lines": [
                            line_value
                        ] if line_value else None,
                        "crop": crop_value,
                        "crop_family": crop_family_value,
                        "variety": variety_value,
                        "client": client_value,
                        "year": int(year_value) if year_value else None,
                        "month": int(month_value) if month_value else None
                    }

                    save_question_intent(
                        memory,
                        question,
                        exact_intent
                    )

                if save_learning_memory(memory):

                    st.success(
                        "Saved. The app will use these corrections on future questions."
                    )

                else:

                    st.warning(
                        "I could not save the learning file. On Streamlit Cloud this can happen if the app folder is read-only."
                    )

# =====================================================
# TITLE
# =====================================================

st.title(
    "🌱 Nursery Intelligence Copilot Enterprise"
)

# =====================================================
# SIDEBAR
# =====================================================

with st.sidebar:

    st.subheader("📁 Data")

    st.write(
        f"Orders Rows: {len(df_orders)}"
    )

    st.write(
        f"Sales Rows: {len(df_sales)}"
    )

    st.write(
        f"Returns Rows: {len(df_returns)}"
    )

# =====================================================
# AI SEARCH
# =====================================================

question = st.text_input(
    "Ask anything about orders, sales, returns, crops, varieties, clients or reps"
)

if question:

    intent = detect_intent(question)

    render_understanding_panel(
        question,
        intent
    )

    # =================================================
    # ACTIVE DATASET
    # =================================================

    if intent["source"] == "sales":

        df_active = df_sales
        source_name = "Sales"

    elif intent["source"] == "returns":

        df_active = df_returns
        source_name = "Returns"

    else:

        df_active = df_orders
        source_name = "Orders"

    st.caption(
        f"Using Dataset: {source_name}"
    )

    if intent["client"]:

        st.caption(
            f"Client: {intent['client']}"
        )

    if intent["crop"]:

        st.caption(
            f"Crop: {intent['crop']}"
        )

    if intent["variety"]:

        st.caption(
            f"Variety: {intent['variety']}"
        )

    if intent.get("lines"):

        st.caption(
            f"Lines: {', '.join(intent['lines'])}"
        )

    elif intent["line"]:

        st.caption(
            f"Line: {intent['line']}"
        )

    # =================================================
    # FILTERED DATA
    # =================================================

    df_temp = apply_filters(
        df_active,
        intent
    )

    # =================================================
    # COMPARE LOGIC
    # =================================================

    if (
        intent["seed"]
        and intent["source"] == "orders"
    ):

        show_seed_forecast(
            question,
            intent
        )

    elif intent["compare"]:

        ql = question.lower()

        compare_items = detect_lines(
            question
        )

        # =============================================
        # SALES VS RETURNS
        # =============================================

        if (
            "sales vs returns" in ql
            or (
                "sales" in ql
                and "returns" in ql
            )
        ):

            sales_filtered = apply_filters(
                df_sales,
                intent
            )

            returns_filtered = apply_filters(
                df_returns,
                intent
            )

            if (
                "lines" in ql
                or (
                    "line" in ql
                    and not intent["line"]
                )
            ):

                sales_by_line = (
                    sales_filtered
                    .groupby("Line")
                    .agg(
                        Sales_Quantity=("Quantity", "sum"),
                        Sales_Amount=("Amount", "sum")
                    )
                )

                returns_by_line = (
                    returns_filtered
                    .groupby("Line")
                    .agg(
                        Returns_Quantity=("Quantity", "sum"),
                        Returns_Amount=("Amount", "sum")
                    )
                )

                compare_df = (
                    sales_by_line
                    .join(
                        returns_by_line,
                        how="outer"
                    )
                    .fillna(0)
                    .reset_index()
                )

                compare_df["Net_Sales"] = (
                    compare_df["Sales_Amount"]
                    - compare_df["Returns_Amount"]
                )

                compare_df = compare_df.sort_values(
                    "Net_Sales",
                    ascending=False
                )

                st.subheader(
                    "Sales vs Returns By Line"
                )

                st.dataframe(
                    compare_df,
                    use_container_width=True
                )

                fig = px.bar(
                    compare_df,
                    x="Line",
                    y=[
                        "Sales_Amount",
                        "Returns_Amount",
                        "Net_Sales"
                    ],
                    barmode="group"
                )

                st.plotly_chart(
                    fig,
                    use_container_width=True
                )

            else:

                sales_amount = (
                    sales_filtered["Amount"]
                    .sum()
                )

                returns_amount = (
                    returns_filtered["Amount"]
                    .sum()
                )

                sales_qty = (
                    sales_filtered["Quantity"]
                    .sum()
                )

                returns_qty = (
                    returns_filtered["Quantity"]
                    .sum()
                )

                net_sales = (
                    sales_amount
                    - returns_amount
                )

                compare_df = pd.DataFrame({

                    "Metric": [

                        "Sales Amount",
                        "Returns Amount",
                        "Net Sales",
                        "Sales Quantity",
                        "Returns Quantity"
                    ],

                    "Value": [

                        sales_amount,
                        returns_amount,
                        net_sales,
                        sales_qty,
                        returns_qty
                    ]
                })

                st.subheader(
                    "Sales vs Returns"
                )

                st.dataframe(compare_df)

                fig = px.bar(
                    compare_df,
                    x="Metric",
                    y="Value"
                )

                st.plotly_chart(
                    fig,
                    use_container_width=True
                )

        # =============================================
        # LINE COMPARISON
        # =============================================

        elif len(compare_items) >= 2:

            results = []

            for item in compare_items:

                temp = df_temp[
                    df_temp["Line"]
                    .astype(str)
                    .map(
                        lambda value: line_exact_matches(
                            value,
                            item
                        )
                    )
                    .fillna(False)
                    .astype(bool)
                ]

                qty_total = (
                    temp["Quantity"]
                    .sum()
                )

                amount_total = (
                    temp["Amount"]
                    .sum()
                )

                results.append({

                    "Line": item,
                    "Quantity": qty_total,
                    "Amount": amount_total
                })

            results_df = pd.DataFrame(results)

            st.subheader(
                "📊 Comparison Results"
            )

            st.dataframe(results_df)

            metric_col = "Quantity"

            if intent["metric"] == "amount":

                metric_col = "Amount"

            fig = px.bar(
                results_df,
                x="Line",
                y=metric_col
            )

            st.plotly_chart(
                fig,
                use_container_width=True
            )

    # =================================================
    # TOP LOGIC
    # =================================================

    elif intent["top"]:

        ql = question.lower()
        group_col = "Line"
        group_label = "Line"
        action_label = "ordered"

        if intent["source"] == "sales":

            action_label = "sold"

        elif intent["source"] == "returns":

            action_label = "returned"

        metric_col = "Quantity"

        if (
            intent["metric"] == "amount"
            and intent["source"] != "orders"
        ):

            metric_col = "Amount"

        if (
            "client" in ql
            or "customer" in ql
            or "store" in ql
        ):

            group_col = "Client Name"
            group_label = "Client"

        if "crop" in ql:

            group_col = "Crop Name"
            group_label = "Crop"

        if "variety" in ql:

            group_col = "Variety"
            group_label = "Variety"

        if "rep" in ql:

            group_col = "Rep"
            group_label = "Rep"

        result = (

            df_temp
            .groupby(group_col)[metric_col]
            .sum()
            .sort_values(
                ascending=False
            )
            .head(10)
        )

        st.subheader(
            f"Top {group_label} {action_label.title()}"
        )

        st.dataframe(result)

        if len(result) > 0:

            top_name = result.index[0]
            top_value = result.iloc[0]

            if metric_col == "Amount":

                st.success(
                    f"{top_name} {action_label} the most: R{top_value:,.2f}"
                )

            else:

                st.success(
                    f"{top_name} {action_label} the most: {top_value:,.0f} units"
                )

            if (
                group_col == "Client Name"
                and (
                    "of what" in ql
                    or "what" in ql
                    or "which line" in ql
                    or "which lines" in ql
                    or "products" in ql
                    or "items" in ql
                )
            ):

                client_rows = df_temp[
                    df_temp["Client Name"]
                    .astype(str)
                    .str.strip()
                    == str(top_name).strip()
                ]

                item_breakdown = (
                    client_rows
                    .groupby("Line")[metric_col]
                    .sum()
                    .sort_values(
                        ascending=False
                    )
                    .head(10)
                )

                st.subheader(
                    f"What {top_name} {action_label}"
                )

                st.dataframe(item_breakdown)

                item_fig = px.bar(
                    x=item_breakdown.index,
                    y=item_breakdown.values
                )

                st.plotly_chart(
                    item_fig,
                    use_container_width=True
                )

        fig = px.bar(
            x=result.index,
            y=result.values
        )

        st.plotly_chart(
            fig,
            use_container_width=True
        )

    # =================================================
    # STANDARD TOTALS
    # =================================================

    else:

        # =============================================
        # ORDERS
        # =============================================

        if intent["source"] == "orders":

            qty_total = (
                df_temp["Quantity"]
                .sum()
            )

            st.success(
                f"""
Total stock ordered:
{qty_total:,.0f} units
"""
            )

        # =============================================
        # SALES / RETURNS
        # =============================================

        elif intent["source"] == "returns":

            returned_qty = (
                df_temp["Quantity"]
                .sum()
            )

            returned_amount = (
                df_temp["Amount"]
                .sum()
            )

            st.success(
                f"""
Returned Quantity:
{returned_qty:,.0f}

Returns Value:
R{returned_amount:,.2f}
"""
            )

        else:

            returns_filtered = apply_filters(
                df_returns,
                intent
            )

            sold_qty = (
                df_temp["Quantity"]
                .sum()
            )

            sold_amount = (
                df_temp["Amount"]
                .sum()
            )

            returned_amount = (
                returns_filtered["Amount"]
                .sum()
            )

            net_sales = (
                sold_amount
                - returned_amount
            )

            st.success(
                f"""
Sold Quantity:
{sold_qty:,.0f}

Sales Revenue:
R{sold_amount:,.2f}

Returns Value:
R{returned_amount:,.2f}

Net Sales:
R{net_sales:,.2f}
"""
            )

    # =================================================
    # MATCHING DATA
    # =================================================

    st.subheader(
        "📋 Matching Data"
    )

    st.dataframe(
        df_temp,
        use_container_width=True
    )

    render_learning_feedback(
        question,
        intent
    )

# =====================================================
# DASHBOARD
# =====================================================

st.header("📊 Dashboard")

with st.sidebar:

    dashboard = st.selectbox(

        "Choose Dashboard",

        [
            "Orders",
            "Sales",
            "Returns",
            "Sowing Planner"
        ]
    )

# =====================================================
# ORDERS DASHBOARD
# =====================================================

if dashboard == "Orders":

    df_orders_dashboard = ensure_standard_columns(
        df_orders.copy()
    )

    col1, col2, col3 = st.columns(3)

    col1.metric(
        "Total Orders",
        len(df_orders_dashboard)
    )

    col2.metric(
        "Total Ordered Qty",
        f"{df_orders_dashboard['Quantity'].sum():,.0f}"
    )

    col3.metric(
        "Recommended Seeds",
        f"{df_orders_dashboard['Seeds Recommended'].sum():,.0f}"
    )

    st.subheader(
        "📦 Orders By Line"
    )

    if len(df_orders_dashboard) > 0:

        orders_lines = (

            df_orders_dashboard
            .groupby("Line")["Quantity"]
            .sum()
            .sort_values(
                ascending=False
            )
        )

        st.dataframe(orders_lines)

        fig = px.bar(
            x=orders_lines.index,
            y=orders_lines.values
        )

        st.plotly_chart(
            fig,
            use_container_width=True
        )

elif dashboard == "Sowing Planner":

    render_sowing_planner_export()

# =====================================================
# SALES DASHBOARD
# =====================================================

elif dashboard == "Sales":

    col1, col2 = st.columns(2)

    col1.metric(
        "Sales Records",
        len(df_sales)
    )

    col2.metric(
        "Sales Revenue",
        f"R{df_sales['Amount'].sum():,.2f}"
    )

    if len(df_sales) > 0:

        st.subheader(
            "💰 Sales By Line"
        )

        sales_lines = (

            df_sales
            .groupby("Line")["Amount"]
            .sum()
            .sort_values(
                ascending=False
            )
        )

        st.dataframe(sales_lines)

        fig = px.bar(
            x=sales_lines.index,
            y=sales_lines.values
        )

        st.plotly_chart(
            fig,
            use_container_width=True
        )

# =====================================================
# RETURNS DASHBOARD
# =====================================================

elif dashboard == "Returns":

    col1, col2 = st.columns(2)

    col1.metric(
        "Returns Records",
        len(df_returns)
    )

    col2.metric(
        "Returns Value",
        f"R{df_returns['Amount'].sum():,.2f}"
    )

    if len(df_returns) > 0:

        st.subheader(
            "🔄 Returns By Line"
        )

        returns_lines = (

            df_returns
            .groupby("Line")["Amount"]
            .sum()
            .sort_values(
                ascending=False
            )
        )

        st.dataframe(returns_lines)

        fig = px.bar(
            x=returns_lines.index,
            y=returns_lines.values
        )

        st.plotly_chart(
            fig,
            use_container_width=True
        )
